"""Excel report generation utilities.

Provides templates and helpers for creating professional Excel reports
with formatting, charts, and data validation.
"""

from typing import List, Dict, Tuple, Optional, Any
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime


class ExcelReport:
    """Builder for creating Excel reports with multiple sheets."""

    def __init__(self, title: str = "Report"):
        """Initialize Excel report.

        Args:
            title: Report title
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Summary"
        self.title = title
        self._setup_styles()

    def _setup_styles(self):
        """Define style templates."""
        self.styles = {
            "title": Font(name="Calibri", size=18, bold=True, color="FFFFFF"),
            "title_fill": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
            "header": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "header_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "subheader": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
            "subheader_fill": PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),
            "metric_label": Font(name="Calibri", size=10, bold=True),
            "metric_value": Font(name="Calibri", size=14, bold=True, color="2E75B6"),
            "critical": Font(name="Calibri", size=10, color="C5504D"),
            "critical_fill": PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid"),
            "high": Font(name="Calibri", size=10, color="ED7D31"),
            "high_fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "medium": Font(name="Calibri", size=10, color="F8B500"),
            "medium_fill": PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),
            "low": Font(name="Calibri", size=10, color="595959"),
            "low_fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "success": Font(name="Calibri", size=10, color="00B050"),
            "success_fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "warning": Font(name="Calibri", size=10, color="FFB347"),
            "warning_fill": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
            "center_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "right_align": Alignment(horizontal="right", vertical="center"),
            "left_align": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            ),
        }

    def add_title(self, title: str, subtitle: str = ""):
        """Add title to current sheet.

        Args:
            title: Main title
            subtitle: Optional subtitle
        """
        self.ws.merge_cells("A1:H1")
        cell = self.ws["A1"]
        cell.value = title
        cell.font = self.styles["title"]
        cell.fill = self.styles["title_fill"]
        cell.alignment = self.styles["center_align"]
        self.ws.row_dimensions[1].height = 30

        if subtitle:
            self.ws.merge_cells("A2:H2")
            cell = self.ws["A2"]
            cell.value = subtitle
            cell.font = Font(name="Calibri", size=11, italic=True)
            cell.alignment = self.styles["center_align"]

    def add_headers(self, headers: List[str], row: int = 1):
        """Add header row.

        Args:
            headers: List of header texts
            row: Row number (1-indexed)
        """
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.styles["header"]
            cell.fill = self.styles["header_fill"]
            cell.alignment = self.styles["center_align"]
            cell.border = self.styles["border"]

    def add_data_table(self, data: List[Dict], start_row: int = 1, start_col: int = 1):
        """Add data table to sheet.

        Args:
            data: List of dictionaries with data
            start_row: Starting row (1-indexed)
            start_col: Starting column (1-indexed)
        """
        if not data:
            return

        # Add headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, start_col):
            cell = self.ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = self.styles["header"]
            cell.fill = self.styles["header_fill"]
            cell.alignment = self.styles["center_align"]
            cell.border = self.styles["border"]

        # Add data rows
        for row_idx, row_data in enumerate(data, start_row + 1):
            for col_idx, header in enumerate(headers, start_col):
                cell = self.ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(header)
                cell.alignment = self.styles["left_align"]
                cell.border = self.styles["border"]

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start_col):
            column_letter = get_column_letter(col_idx)
            self.ws.column_dimensions[column_letter].width = len(header) + 5

    def add_metrics_grid(self, metrics: Dict[str, Any], rows: int = 2, cols: int = 2):
        """Add key metrics grid to sheet.

        Args:
            metrics: Dictionary of {metric_name: metric_value}
            rows: Number of rows in grid
            cols: Number of columns in grid
        """
        start_row = self.ws.max_row + 2

        for idx, (name, value) in enumerate(metrics.items()):
            if idx >= rows * cols:
                break

            row = start_row + (idx // cols) * 3
            col = (idx % cols) * 4 + 1

            # Metric name
            name_cell = self.ws.cell(row=row, column=col)
            name_cell.value = name
            name_cell.font = self.styles["metric_label"]

            # Metric value
            value_cell = self.ws.cell(row=row + 1, column=col)
            value_cell.value = value
            value_cell.font = self.styles["metric_value"]

            self.ws.merge_cells(
                f"{get_column_letter(col)}{row}:{get_column_letter(col+2)}{row}"
            )
            self.ws.merge_cells(
                f"{get_column_letter(col)}{row+1}:{get_column_letter(col+2)}{row+1}"
            )

    def add_severity_summary(self, summary: Dict[str, int]):
        """Add severity summary (critical, high, medium, low).

        Args:
            summary: Dictionary with severity counts
        """
        start_row = self.ws.max_row + 2

        severity_order = ["critical", "high", "medium", "low"]
        colors = {
            "critical": self.styles["critical_fill"],
            "high": self.styles["high_fill"],
            "medium": self.styles["medium_fill"],
            "low": self.styles["low_fill"],
        }

        # Header
        self.ws.merge_cells(f"A{start_row}:D{start_row}")
        cell = self.ws[f"A{start_row}"]
        cell.value = "Findings by Severity"
        cell.font = self.styles["subheader"]
        cell.fill = self.styles["subheader_fill"]
        cell.alignment = self.styles["center_align"]

        # Severity rows
        for idx, severity in enumerate(severity_order):
            row = start_row + 1 + idx
            count = summary.get(severity, 0)

            label_cell = self.ws[f"A{row}"]
            label_cell.value = severity.capitalize()
            label_cell.fill = colors.get(severity)

            count_cell = self.ws[f"B{row}"]
            count_cell.value = count
            count_cell.fill = colors.get(severity)

    def add_sheet(self, name: str):
        """Add a new sheet to the workbook.

        Args:
            name: Sheet name

        Returns:
            The new worksheet
        """
        ws = self.wb.create_sheet(name)
        self.ws = ws
        return ws

    def freeze_panes(self, row: int = 1, col: int = 1):
        """Freeze panes at specified row/column.

        Args:
            row: Row to freeze at
            col: Column to freeze at
        """
        cell = self.ws.cell(row=row + 1, column=col + 1)
        self.ws.freeze_panes = cell

    def set_column_width(self, column: str, width: float):
        """Set column width.

        Args:
            column: Column letter (e.g. "A")
            width: Width in characters
        """
        self.ws.column_dimensions[column].width = width

    def set_row_height(self, row: int, height: float):
        """Set row height.

        Args:
            row: Row number
            height: Height in points
        """
        self.ws.row_dimensions[row].height = height

    def add_image(self, image_buffer: BytesIO, position: str = "A1", width: int = 400, height: int = 300):
        """Add image to sheet.

        Args:
            image_buffer: BytesIO buffer with image data
            position: Cell position (e.g. "A1")
            width: Image width in pixels
            height: Image height in pixels
        """
        img = XLImage(image_buffer)
        img.width = width
        img.height = height
        self.ws.add_image(img, position)

    def add_footer(self, text: str = ""):
        """Add footer to sheet.

        Args:
            text: Footer text (default: timestamp)
        """
        if not text:
            text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_LETTER
        self.ws.page_margins.left = 0.5
        self.ws.page_margins.right = 0.5
        self.ws.page_margins.top = 0.75
        self.ws.page_margins.bottom = 0.75

        self.ws.footer.center.text = text

    def save(self, filepath: str):
        """Save workbook to file.

        Args:
            filepath: Output file path

        Returns:
            File path
        """
        self.wb.save(filepath)
        return filepath

    def to_bytes(self) -> bytes:
        """Convert workbook to bytes.

        Returns:
            Bytes representation of workbook
        """
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
