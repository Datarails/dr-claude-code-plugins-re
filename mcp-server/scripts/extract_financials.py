#!/usr/bin/env python3
"""
Datarails Financial Data Extraction Script

Extracts P&L and KPI data from Finance OS and generates Excel workbook.
Supports client-specific profiles for different table structures.

Usage:
    uv --directory mcp-server run python scripts/extract_financials.py --year 2025
    uv --directory mcp-server run python scripts/extract_financials.py --year 2025 --env app
    uv --directory mcp-server run python scripts/extract_financials.py --year 2025 --profile config/client-profiles/app.json
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# Add mcp-server/src to path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import httpx
import asyncio
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from datarails_mcp.auth import get_auth


# Default profile (backward compatibility with original hardcoded values)
DEFAULT_PROFILE = {
    "version": "1.0",
    "environment": "app",
    "tables": {
        "financials": {
            "id": "16528",
            "name": "Financials Cube",
            "purpose": "P&L and Balance Sheet"
        },
        "kpis": {
            "id": "34298",
            "name": "KPI Metrics",
            "purpose": "ARR, Churn, LTV"
        }
    },
    "field_mappings": {
        "amount": "Amount",
        "scenario": "Scenario",
        "year": "System_Year",
        "date": "Reporting Date",
        "account_l0": "DR_ACC_L0",
        "account_l1": "DR_ACC_L1",
        "account_l2": "DR_ACC_L2",
        "kpi_name": "KPI",
        "kpi_value": "value",
        "quarter": "Quarter & Year"
    },
    "account_hierarchy": {
        "pnl_filter": "P&L",
        "balance_sheet_filter": "Balance Sheet",
        "revenue": "REVENUE",
        "cogs": "Cost of Good sold",
        "opex": "Operating Expense",
        "financial_expense": "Financial Expenses",
        "other_expense": "Other Expense"
    },
    "kpi_definitions": {
        "revenue": "Revenue",
        "gross_profit": "Gross Profit",
        "arr_opening": "ARR Opening Balance",
        "new_arr": "New ARR",
        "net_new_arr": "Net New ARR",
        "churn_dollar": "Churn $",
        "churn_pct": "Churn %",
        "ltv": "LTV",
        "ltv_cac": "LTV / CAC"
    },
    "scenarios": ["Actuals", "Budget", "Forecast"],
    "years_available": ["2023", "2024", "2025", "2026"]
}


def load_profile(profile_path: str = None, env: str = None) -> dict:
    """Load client profile from file or use defaults.

    Args:
        profile_path: Direct path to profile JSON file
        env: Environment name to load profile for (looks in config/client-profiles/<env>.json)

    Returns:
        Profile dictionary
    """
    # Try explicit profile path first
    if profile_path:
        path = Path(profile_path)
        if path.exists():
            print(f"  Loading profile: {profile_path}")
            with open(path) as f:
                return json.load(f)
        else:
            print(f"  Warning: Profile not found at {profile_path}, using defaults")
            return DEFAULT_PROFILE

    # Try environment-based profile
    if env:
        # Look relative to script location (mcp-server/scripts/) -> project root
        project_root = Path(__file__).parent.parent.parent
        profile_path = project_root / "config" / "client-profiles" / f"{env}.json"

        if profile_path.exists():
            print(f"  Loading profile: {profile_path}")
            with open(profile_path) as f:
                return json.load(f)
        else:
            print(f"  No profile found for '{env}' environment")
            print(f"  Run '/dr-learn --env {env}' to create one, or using defaults")
            return DEFAULT_PROFILE

    # Fall back to defaults
    print("  Using default profile (original hardcoded values)")
    return DEFAULT_PROFILE


class FinancialExtractor:
    """Extract financial data from Datarails Finance OS."""

    def __init__(self, year: int, scenario: str = "Actuals", profile: dict = None):
        self.year = year
        self.scenario = scenario
        self.profile = profile or DEFAULT_PROFILE
        self.auth = get_auth()
        self.base_url = self.auth.base_url

        # Extract table IDs from profile
        self.financials_table = self.profile["tables"]["financials"]["id"]
        self.kpis_table = self.profile["tables"].get("kpis", {}).get("id")

        # Extract field mappings
        self.fields = self.profile["field_mappings"]

        # Extract account hierarchy
        self.accounts = self.profile.get("account_hierarchy", {})

        # Extract KPI definitions
        self.kpis = self.profile.get("kpi_definitions", {})

    def ensure_auth(self):
        """Ensure we have valid authentication."""
        asyncio.run(self.auth.ensure_valid_token())

    def get_headers(self):
        """Get authenticated headers."""
        return self.auth.get_headers()

    def fetch_pages(self, table_id: str, filters: list, max_rows: int = 100000) -> list:
        """Fetch data with pagination and token refresh."""
        all_data = []
        offset = 0
        url = f'{self.base_url}/finance-os/api/tables/v1/{table_id}/data'

        print(f"  Fetching from table {table_id}...")

        while len(all_data) < max_rows:
            # Re-auth every 20K rows
            if offset > 0 and offset % 20000 == 0:
                print(f"    Refreshing token at {len(all_data)} rows...")
                self.ensure_auth()

            try:
                resp = httpx.post(
                    url,
                    headers=self.get_headers(),
                    json={
                        'filters': filters,
                        'limit': 500,
                        'offset': offset,
                    },
                    timeout=60.0
                )

                if resp.status_code == 401:
                    print("    Token expired, refreshing...")
                    self.ensure_auth()
                    resp = httpx.post(
                        url,
                        headers=self.get_headers(),
                        json={
                            'filters': filters,
                            'limit': 500,
                            'offset': offset,
                        },
                        timeout=60.0
                    )

                if resp.status_code != 200:
                    print(f"    Error at offset {offset}: {resp.status_code}")
                    break

                page = resp.json().get('data', [])
                if not page:
                    break

                all_data.extend(page)

                if len(all_data) % 10000 == 0:
                    print(f"    Fetched {len(all_data)} rows...")

                if len(page) < 500:
                    break

                offset += 500

            except Exception as e:
                print(f"    Error: {e}")
                break

        print(f"    Total: {len(all_data)} rows")
        return all_data

    def extract_pl(self) -> dict:
        """Extract P&L data aggregated by month and account."""
        print(f"\nExtracting {self.year} P&L ({self.scenario})...")

        # Build filters using profile field names
        scenario_field = self.fields.get("scenario", "Scenario")
        account_l0_field = self.fields.get("account_l0", "DR_ACC_L0")
        year_field = self.fields.get("year", "System_Year")
        pnl_filter_value = self.accounts.get("pnl_filter", "P&L")

        data = self.fetch_pages(
            table_id=self.financials_table,
            filters=[
                {"name": scenario_field, "values": [self.scenario], "is_excluded": False},
                {"name": account_l0_field, "values": [pnl_filter_value], "is_excluded": False},
                {"name": year_field, "values": [str(self.year)], "is_excluded": False},
            ]
        )

        # Get field names from profile
        date_field = self.fields.get("date", "Reporting Date")
        account_l1_field = self.fields.get("account_l1", "DR_ACC_L1")
        amount_field = self.fields.get("amount", "Amount")

        # Aggregate by L1 and month
        pl_agg = defaultdict(lambda: defaultdict(float))
        months = set()

        for row in data:
            rd = row.get(date_field)
            if rd:
                month = datetime.fromtimestamp(rd).strftime('%Y-%m')
                l1 = row.get(account_l1_field, 'Unknown')
                amt = row.get(amount_field, 0) or 0
                pl_agg[l1][month] += amt
                months.add(month)

        print(f"  Months covered: {sorted(months)}")
        print(f"  Account categories: {list(pl_agg.keys())}")

        return dict(pl_agg), sorted(months)

    def extract_kpis(self) -> dict:
        """Extract KPI data by quarter."""
        if not self.kpis_table:
            print(f"\nNo KPI table configured, skipping KPI extraction")
            return {}, []

        print(f"\nExtracting {self.year} KPIs ({self.scenario})...")

        scenario_field = self.fields.get("scenario", "Scenario")

        data = self.fetch_pages(
            table_id=self.kpis_table,
            filters=[
                {"name": scenario_field, "values": [self.scenario], "is_excluded": False},
            ],
            max_rows=10000
        )

        # Get field names from profile
        kpi_name_field = self.fields.get("kpi_name", "KPI")
        kpi_value_field = self.fields.get("kpi_value", "value")
        quarter_field = self.fields.get("quarter", "Quarter & Year")
        date_field = self.fields.get("date", "Reporting Date")

        # Extract metrics for target year
        metrics = defaultdict(dict)
        quarters = set()

        for row in data:
            kpi = row.get(kpi_name_field, '')
            val = row.get(kpi_value_field)
            qtr = row.get(quarter_field, '')
            rd = row.get(date_field)

            if rd and val is not None:
                year = datetime.fromtimestamp(rd).year
                if year == self.year:
                    metrics[kpi][qtr] = val
                    quarters.add(qtr)

        print(f"  Quarters: {sorted(quarters)}")
        print(f"  KPIs found: {len(metrics)}")

        return dict(metrics), sorted(quarters)

    def generate_excel(self, pl_data: dict, pl_months: list, kpi_data: dict, kpi_quarters: list, output_path: str):
        """Generate Excel workbook."""
        print(f"\nGenerating Excel...")

        wb = Workbook()
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        money_format = '#,##0.00'
        pct_format = '0.00%'

        # Get account names from profile
        revenue_name = self.accounts.get("revenue", "REVENUE")
        cogs_name = self.accounts.get("cogs", "Cost of Good sold")
        opex_name = self.accounts.get("opex", "Operating Expense")
        fin_expense_name = self.accounts.get("financial_expense", "Financial Expenses")
        other_expense_name = self.accounts.get("other_expense", "Other Expense")

        # Get KPI names from profile
        revenue_kpi = self.kpis.get("revenue", "Revenue")
        gross_profit_kpi = self.kpis.get("gross_profit", "Gross Profit")
        arr_opening_kpi = self.kpis.get("arr_opening", "ARR Opening Balance")
        net_new_arr_kpi = self.kpis.get("net_new_arr", "Net New ARR")
        new_arr_kpi = self.kpis.get("new_arr", "New ARR")
        churn_dollar_kpi = self.kpis.get("churn_dollar", "Churn $")
        churn_pct_kpi = self.kpis.get("churn_pct", "Churn %")
        ltv_kpi = self.kpis.get("ltv", "LTV")
        ltv_cac_kpi = self.kpis.get("ltv_cac", "LTV / CAC")
        opex_kpi = self.kpis.get("opex", "Operating Expense")

        # === Summary Sheet ===
        summary = wb.active
        summary.title = 'Summary'
        summary['A1'] = f'{self.year} Financial Summary'
        summary['A1'].font = Font(bold=True, size=16)
        summary['A2'] = f'Scenario: {self.scenario}'
        summary['A3'] = f'Extracted: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        summary['A4'] = f'Profile: {self.profile.get("environment", "default")}'

        # P&L Summary
        summary['A6'] = 'P&L Summary'
        summary['A6'].font = Font(bold=True, size=12)
        row = 7
        for l1 in [revenue_name, cogs_name, opex_name, fin_expense_name, other_expense_name]:
            if l1 in pl_data:
                total = sum(pl_data[l1].values())
                summary[f'A{row}'] = l1
                summary[f'B{row}'] = total
                summary[f'B{row}'].number_format = money_format
                row += 1

        # KPI Summary
        if kpi_data:
            summary[f'A{row + 1}'] = 'KPI Summary'
            summary[f'A{row + 1}'].font = Font(bold=True, size=12)
            row += 2
            key_kpis = [revenue_kpi, gross_profit_kpi, arr_opening_kpi, net_new_arr_kpi, new_arr_kpi]
            for kpi in key_kpis:
                if kpi in kpi_data:
                    if 'Opening' in kpi or 'Balance' in kpi:
                        # Show latest quarter
                        val = kpi_data[kpi].get(kpi_quarters[-1], 0) if kpi_quarters else 0
                    else:
                        val = sum(kpi_data[kpi].values())
                    summary[f'A{row}'] = kpi
                    summary[f'B{row}'] = val
                    summary[f'B{row}'].number_format = money_format
                    row += 1

        # === P&L Sheet ===
        pl_sheet = wb.create_sheet('P&L')
        headers = ['Account'] + pl_months + ['Total']
        for col, h in enumerate(headers, 1):
            cell = pl_sheet.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        row_num = 2
        for l1 in sorted(pl_data.keys()):
            pl_sheet.cell(row=row_num, column=1, value=l1)
            total = 0
            for col, month in enumerate(pl_months, 2):
                val = pl_data[l1].get(month, 0)
                pl_sheet.cell(row=row_num, column=col, value=val).number_format = money_format
                total += val
            pl_sheet.cell(row=row_num, column=len(pl_months) + 2, value=total).number_format = money_format
            row_num += 1

        # Auto-width
        for col in range(1, len(headers) + 1):
            pl_sheet.column_dimensions[get_column_letter(col)].width = 15
        pl_sheet.column_dimensions['A'].width = 25

        # === KPIs Sheet ===
        if kpi_data and kpi_quarters:
            kpi_sheet = wb.create_sheet('KPIs')
            headers = ['KPI'] + kpi_quarters + ['Total/Latest']
            for col, h in enumerate(headers, 1):
                cell = kpi_sheet.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            key_kpis_full = [
                revenue_kpi, gross_profit_kpi, opex_kpi,
                arr_opening_kpi, new_arr_kpi, net_new_arr_kpi,
                churn_dollar_kpi, churn_pct_kpi, ltv_kpi, ltv_cac_kpi
            ]

            row_num = 2
            for kpi in key_kpis_full:
                if kpi in kpi_data:
                    kpi_sheet.cell(row=row_num, column=1, value=kpi)
                    total = 0
                    for col, qtr in enumerate(kpi_quarters, 2):
                        val = kpi_data[kpi].get(qtr, 0) or 0
                        kpi_sheet.cell(row=row_num, column=col, value=val).number_format = money_format
                        total += val

                    # Total or Latest depending on KPI type
                    if 'Balance' in kpi or 'LTV' in kpi or '%' in kpi:
                        final_val = kpi_data[kpi].get(kpi_quarters[-1], 0) if kpi_quarters else 0
                    else:
                        final_val = total
                    kpi_sheet.cell(row=row_num, column=len(kpi_quarters) + 2, value=final_val).number_format = money_format
                    row_num += 1

            for col in range(1, len(headers) + 1):
                kpi_sheet.column_dimensions[get_column_letter(col)].width = 18
            kpi_sheet.column_dimensions['A'].width = 25

        # === Validation Sheet ===
        val_sheet = wb.create_sheet('Validation')
        val_sheet['A1'] = 'Data Validation'
        val_sheet['A1'].font = Font(bold=True, size=14)

        val_sheet['A3'] = 'Check'
        val_sheet['B3'] = 'P&L Value'
        val_sheet['C3'] = 'KPI Value'
        val_sheet['D3'] = 'Difference'
        val_sheet['E3'] = 'Status'
        for col in range(1, 6):
            val_sheet.cell(row=3, column=col).font = header_font
            val_sheet.cell(row=3, column=col).fill = header_fill

        # Revenue check
        pl_revenue = sum(pl_data.get(revenue_name, {}).values())
        kpi_revenue = sum(kpi_data.get(revenue_kpi, {}).values()) if kpi_data else 0
        diff = abs(pl_revenue - kpi_revenue)
        status = 'PASS' if diff < 10000 else 'CHECK'

        val_sheet['A4'] = 'Revenue Match'
        val_sheet['B4'] = pl_revenue
        val_sheet['C4'] = kpi_revenue
        val_sheet['D4'] = diff
        val_sheet['E4'] = status
        val_sheet['B4'].number_format = money_format
        val_sheet['C4'].number_format = money_format
        val_sheet['D4'].number_format = money_format

        val_sheet['A5'] = 'Months Covered'
        val_sheet['B5'] = len(pl_months)
        val_sheet['E5'] = 'PASS' if len(pl_months) == 12 else 'CHECK'

        val_sheet['A6'] = 'Quarters Covered'
        val_sheet['C6'] = len(kpi_quarters) if kpi_quarters else 0
        val_sheet['E6'] = 'PASS' if len(kpi_quarters) == 4 else 'CHECK' if kpi_quarters else 'N/A'

        # Profile info
        val_sheet['A8'] = 'Profile Information'
        val_sheet['A8'].font = Font(bold=True)
        val_sheet['A9'] = 'Environment'
        val_sheet['B9'] = self.profile.get('environment', 'default')
        val_sheet['A10'] = 'Financials Table'
        val_sheet['B10'] = f"{self.profile['tables']['financials']['name']} (ID: {self.financials_table})"
        if self.kpis_table:
            val_sheet['A11'] = 'KPI Table'
            val_sheet['B11'] = f"{self.profile['tables']['kpis']['name']} (ID: {self.kpis_table})"

        # Save
        wb.save(output_path)
        print(f"  Saved: {output_path}")

    def run(self, output_path: str = None):
        """Run the full extraction."""
        print("=" * 60)
        print(f"Datarails Financial Extraction - {self.year} {self.scenario}")
        print(f"Profile: {self.profile.get('environment', 'default')}")
        print("=" * 60)

        # Ensure authenticated
        print("\nAuthenticating...")
        self.ensure_auth()
        print(f"  Connected to: {self.base_url}")

        # Extract data
        pl_data, pl_months = self.extract_pl()
        kpi_data, kpi_quarters = self.extract_kpis()

        # Generate output path
        if not output_path:
            output_path = f"tmp/Financial_Extract_{self.year}.xlsx"

        # Ensure tmp directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        # Generate Excel
        self.generate_excel(pl_data, pl_months, kpi_data, kpi_quarters, output_path)

        # Print summary
        print("\n" + "=" * 60)
        print("SUMMARY")
        print("=" * 60)

        # Get names from profile
        revenue_name = self.accounts.get("revenue", "REVENUE")
        cogs_name = self.accounts.get("cogs", "Cost of Good sold")
        opex_name = self.accounts.get("opex", "Operating Expense")
        revenue_kpi = self.kpis.get("revenue", "Revenue")
        arr_opening_kpi = self.kpis.get("arr_opening", "ARR Opening Balance")
        net_new_arr_kpi = self.kpis.get("net_new_arr", "Net New ARR")

        print(f"\nP&L ({len(pl_months)} months):")
        for l1 in [revenue_name, cogs_name, opex_name]:
            if l1 in pl_data:
                print(f"  {l1}: ${sum(pl_data[l1].values()):,.2f}")

        if kpi_data:
            print(f"\nKPIs ({len(kpi_quarters)} quarters):")
            for kpi in [revenue_kpi, arr_opening_kpi, net_new_arr_kpi]:
                if kpi in kpi_data:
                    if 'Opening' in kpi:
                        val = kpi_data[kpi].get(kpi_quarters[-1], 0) if kpi_quarters else 0
                        print(f"  {kpi} (Q4): ${val:,.2f}")
                    else:
                        print(f"  {kpi}: ${sum(kpi_data[kpi].values()):,.2f}")

        print(f"\nOutput: {output_path}")
        print("=" * 60)


def main():
    parser = argparse.ArgumentParser(description='Extract Datarails financial data to Excel')
    parser.add_argument('--year', type=int, default=datetime.now().year,
                        help='Calendar year to extract (default: current year)')
    parser.add_argument('--scenario', type=str, default='Actuals',
                        help='Scenario to extract (default: Actuals)')
    parser.add_argument('--output', type=str, default=None,
                        help='Output file path (default: tmp/Financial_Extract_YYYY.xlsx)')
    parser.add_argument('--profile', type=str, default=None,
                        help='Path to client profile JSON file')
    parser.add_argument('--env', type=str, default=None,
                        help='Environment name to load profile for (dev, demo, testapp, app)')

    args = parser.parse_args()

    # Load profile
    profile = load_profile(profile_path=args.profile, env=args.env)

    extractor = FinancialExtractor(year=args.year, scenario=args.scenario, profile=profile)
    extractor.run(output_path=args.output)


if __name__ == '__main__':
    main()
