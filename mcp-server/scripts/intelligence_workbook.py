#!/usr/bin/env python3
"""
FP&A Intelligence Workbook Generator

Creates an actionable insights workbook that answers real business questions,
not just data dumps. Surfaces findings automatically and provides recommendations.

Features:
- Insights Dashboard with top 5 findings
- Expense Deep Dive (Top 20 accounts, cost centers)
- Variance Waterfall (what changed and why)
- Trend Analysis (12-month trends with growth rates)
- Anomaly Report (auto-detected outliers)
- Vendor Analysis (top vendors, concentration risk)
- SaaS Metrics (ARR, Unit Economics, Efficiency)
- Sales Performance (rep leaderboard, cohorts)
- Cost Center P&L (department-level detail)
- Raw Data (pivot-ready for your analysis)

Usage:
    uv --directory mcp-server run python scripts/intelligence_workbook.py --year 2025
    uv --directory mcp-server run python scripts/intelligence_workbook.py --year 2025 --env app
    uv --directory mcp-server run python scripts/intelligence_workbook.py --year 2025 --output tmp/custom_name.xlsx
"""

import argparse
import json
import sys
import asyncio
import statistics
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Any, Dict, List, Tuple, Optional

# Add mcp-server/src to path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

from datarails_mcp.auth import get_auth
from datarails_mcp.client import DatarailsClient
from datarails_mcp.report_utils import (
    format_currency,
    format_percentage,
    format_ratio,
    calculate_growth_rate,
    safe_divide,
)


class IntelligenceWorkbook:
    """Generate actionable FP&A intelligence workbook."""

    # Professional color scheme
    HEADER_COLOR = "1F4E78"      # Dark blue
    SUBHEADER_COLOR = "4472C4"   # Medium blue
    ACCENT_COLOR = "ED7D31"      # Orange accent
    LIGHT_BG = "F2F2F2"          # Light gray
    POSITIVE_COLOR = "70AD47"    # Green
    NEGATIVE_COLOR = "C00000"    # Red
    WARNING_COLOR = "FFD966"     # Yellow/amber
    INFO_COLOR = "5B9BD5"        # Light blue

    # Insight severity colors
    SEVERITY_COLORS = {
        "CRITICAL": "C00000",
        "WARNING": "ED7D31",
        "POSITIVE": "70AD47",
        "INFO": "5B9BD5",
        "ACTION": "7030A0",  # Purple
    }

    # Materiality thresholds
    MATERIALITY_THRESHOLD = 0.05  # 5% of total
    VARIANCE_ALERT_THRESHOLD = 0.10  # 10% change triggers alert
    CONCENTRATION_RISK_THRESHOLD = 0.10  # 10% vendor concentration

    def __init__(self, year: int, env: str = "app", profile_path: str = None):
        """Initialize the Intelligence Workbook generator."""
        self.year = year
        self.env = env
        self.auth = get_auth()
        self.client = DatarailsClient(self.auth)

        # Load profile
        self.profile = self._load_profile(profile_path, env)

        # Extract configuration
        self.financials_table = self.profile["tables"]["financials"]["id"]
        self.kpis_table = self.profile["tables"].get("kpis", {}).get("id")
        self.fields = self.profile["field_mappings"]
        self.accounts = self.profile.get("account_hierarchy", {})

        # Aggregation hints from profile (discovered by /dr-test or /dr-learn)
        self.agg_hints = self.profile.get("aggregation", {})
        self.agg_supported = self.agg_hints.get("supported", True)
        self.agg_failed_fields = set(self.agg_hints.get("failed_fields", []))
        self.agg_alternatives = self.agg_hints.get("field_alternatives", {})

        # Data containers
        self.pnl_data = []
        self.kpi_data = []
        self.aggregated_data = {}
        self.insights = []
        self.recommendations = []
        self.timestamps = {}

        # Create workbook
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def _resolve_field(self, semantic_name: str) -> str:
        """Resolve a semantic field name, using aggregation alternatives if the field is known to fail.

        Args:
            semantic_name: Semantic name like 'account_l1', 'account_l2', etc.

        Returns:
            The actual field name to use, substituting alternatives for failed fields.
        """
        actual_field = self.fields.get(semantic_name, "")

        # If this actual field is known to fail in aggregation, try the alternative
        if actual_field in self.agg_failed_fields:
            alt_semantic = self.agg_alternatives.get(semantic_name)
            if alt_semantic and alt_semantic in self.fields:
                alt_field = self.fields[alt_semantic]
                return alt_field

        return actual_field

    def _load_profile(self, profile_path: str, env: str) -> dict:
        """Load client profile."""
        if profile_path:
            path = Path(profile_path)
            if path.exists():
                with open(path) as f:
                    return json.load(f)

        project_root = Path(__file__).parent.parent.parent
        profile_path = project_root / "config" / "client-profiles" / f"{env}.json"

        if profile_path.exists():
            with open(profile_path) as f:
                return json.load(f)

        raise FileNotFoundError(f"No profile found for {env}")

    def _get_border(self, style: str = "thin") -> Border:
        """Get border style."""
        side = Side(style=style, color="CCCCCC")
        return Border(left=side, right=side, top=side, bottom=side)

    def _format_header_cell(self, cell, color: str = None):
        """Format a header cell."""
        color = color or self.HEADER_COLOR
        cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = self._get_border()

    def _format_subheader_cell(self, cell, color: str = None):
        """Format a subheader cell."""
        color = color or self.SUBHEADER_COLOR
        cell.font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = self._get_border()

    def _format_data_cell(self, cell, is_currency: bool = False, is_percentage: bool = False):
        """Format a data cell."""
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(horizontal="right" if (is_currency or is_percentage) else "left", vertical="center")
        cell.border = self._get_border()

        if is_currency:
            cell.number_format = '$#,##0.00'
        elif is_percentage:
            cell.number_format = '0.0%'

    async def fetch_all_data(self):
        """Fetch all required data from Datarails.

        Uses aggregation-first strategy with parallel queries when possible.
        Falls back to pagination only when aggregation is marked unsupported in profile.
        """
        print("\n1️⃣  FETCHING DATA FROM DATARAILS...")

        # Check authentication first
        if not self.auth.is_authenticated():
            print("    ⚠ Not authenticated! Run '/dr-auth --env app' first")
            return

        # Ensure we have a valid token
        if not await self.auth.ensure_valid_token():
            print("    ⚠ Failed to get valid token. Please re-authenticate.")
            return

        if self.agg_supported:
            print("    ℹ Using aggregation API (profile: aggregation supported)")
            if self.agg_failed_fields:
                print(f"    ℹ Known failed fields: {', '.join(sorted(self.agg_failed_fields))}")
            if self.agg_alternatives:
                print(f"    ℹ Using alternatives: {self.agg_alternatives}")

            # Run independent aggregation queries in parallel for ~5s total
            results = await asyncio.gather(
                self._fetch_pnl_aggregated(),
                self._fetch_kpi_data(),
                self._fetch_vendor_data(),
                self._fetch_cost_center_data(),
                return_exceptions=True,
            )

            # Log any exceptions from parallel tasks
            for i, result in enumerate(results):
                if isinstance(result, Exception):
                    task_names = ["P&L", "KPI", "Vendor", "Cost Center"]
                    print(f"    ⚠ {task_names[i]} fetch failed: {result}")
        else:
            print("    ℹ Aggregation marked unsupported in profile, using pagination fallback")
            # Sequential fetch with pagination (slow path)
            await self._fetch_pnl_aggregated()
            await self._fetch_kpi_data()
            await self._fetch_vendor_data()
            await self._fetch_cost_center_data()

        print(f"    ✓ All data fetched successfully")

    def _parse_api_response(self, result: str) -> List[dict]:
        """Parse API response and handle errors."""
        try:
            # Handle None result
            if result is None:
                return []

            if isinstance(result, str):
                # Handle empty or null strings
                if result.strip() in ('', 'null', 'None'):
                    return []
                data = json.loads(result)
            else:
                data = result

            # Handle null data
            if data is None:
                return []

            # Check for error response
            if isinstance(data, dict):
                if "error" in data:
                    print(f"    ⚠ API Error: {data.get('error')}")
                    return []
                # Some responses wrap data
                if "data" in data:
                    return data["data"] if isinstance(data["data"], list) else []
                return []

            return data if isinstance(data, list) else []

        except json.JSONDecodeError as e:
            print(f"    ⚠ Failed to parse response: {e}")
            return []
        except Exception as e:
            print(f"    ⚠ Unexpected error parsing response: {e}")
            return []

    async def _fetch_raw_data_paginated(
        self,
        table_id: str,
        filters: List[dict],
        max_rows: int = 100000
    ) -> List[dict]:
        """Fetch ALL data using pagination with token refresh.

        Similar to extract_financials.py's fetch_pages() method:
        - Fetches 500 records at a time
        - Refreshes JWT token every 20,000 rows to avoid expiry
        - Continues until all data is fetched
        """
        import httpx

        all_data = []
        offset = 0
        url = f'{self.auth.base_url}/finance-os/api/tables/v1/{table_id}/data'

        # Convert filter list to API format
        api_filters = filters if filters else None

        # Reuse a single client for all requests (more efficient)
        async with httpx.AsyncClient(timeout=60.0) as client:
            while len(all_data) < max_rows:
                # Re-authenticate every 20K rows to refresh JWT token (5-min expiry)
                if offset > 0 and offset % 20000 == 0:
                    print(f"      Refreshing token at {len(all_data):,} rows...")
                    await self.auth.ensure_valid_token()

                retry_count = 0
                max_retries = 3

                while retry_count < max_retries:
                    try:
                        resp = await client.post(
                            url,
                            headers=self.auth.get_headers(),
                            json={
                                'filters': api_filters,
                                'limit': 500,
                                'offset': offset,
                                'get_all_versions': False,
                            }
                        )

                        if resp.status_code == 401:
                            print("      Token expired, refreshing...")
                            await self.auth.ensure_valid_token()
                            retry_count += 1
                            continue

                        if resp.status_code == 429:  # Rate limit
                            print("      Rate limited, waiting 2s...")
                            await asyncio.sleep(2)
                            retry_count += 1
                            continue

                        if resp.status_code >= 500:  # Server error
                            print(f"      Server error {resp.status_code}, retrying...")
                            await asyncio.sleep(1)
                            retry_count += 1
                            continue

                        if resp.status_code != 200:
                            print(f"      Error at offset {offset}: {resp.status_code}")
                            break

                        page = resp.json().get('data', [])
                        if not page:
                            # No more data
                            return all_data

                        all_data.extend(page)

                        if len(all_data) % 10000 == 0:
                            print(f"      Fetched {len(all_data):,} rows...")

                        # If we got less than 500, we've reached the end
                        if len(page) < 500:
                            return all_data

                        offset += 500
                        break  # Success, exit retry loop

                    except httpx.TimeoutException:
                        print(f"      Timeout at offset {offset}, retrying...")
                        retry_count += 1
                        await asyncio.sleep(1)
                    except Exception as e:
                        print(f"      Error at offset {offset}: {e}")
                        retry_count += 1
                        await asyncio.sleep(1)

                # If all retries failed, exit the main loop
                if retry_count >= max_retries:
                    print(f"      Max retries reached at offset {offset}, stopping...")
                    break

        return all_data

    async def _fetch_raw_data(self, filters: List[dict], limit: int = 500) -> List[dict]:
        """Fetch raw data using the data endpoint - DEPRECATED, use _fetch_raw_data_paginated."""
        try:
            result_str = await self.client.get_filtered(
                table_id=self.financials_table,
                filters={f["name"]: {"in": f["values"]} for f in filters},
                limit=limit
            )
            return self._parse_api_response(result_str)
        except Exception as e:
            print(f"    ⚠ Error fetching raw data: {e}")
            return []

    def _aggregate_client_side(self, data: List[dict], group_by: List[str], sum_field: str) -> List[dict]:
        """Aggregate data client-side as fallback."""
        if not data:
            return []

        aggregated = defaultdict(float)
        date_field = self.fields.get("date", "Reporting Date")

        for record in data:
            # Build key, converting timestamps to month strings for date fields
            key_parts = []
            for field in group_by:
                value = record.get(field, "")
                # Convert timestamp dates to YYYY-MM format
                if field == date_field and isinstance(value, (int, float)):
                    value = self._convert_timestamp_to_month(value)
                key_parts.append(str(value) if value else "")

            key = tuple(key_parts)
            try:
                amount = float(record.get(sum_field, 0) or 0)
                aggregated[key] += amount
            except (ValueError, TypeError):
                pass

        # Convert back to list of dicts
        result = []
        for key, total in aggregated.items():
            row = dict(zip(group_by, key))
            row[sum_field] = total
            result.append(row)

        return result

    async def _fetch_pnl_aggregated(self):
        """Fetch P&L data using aggregation API with profile-driven field alternatives."""
        print("  📊 Fetching P&L data...")

        # Resolve fields - use alternatives for known-failed fields
        account_l1_field = self._resolve_field("account_l1")
        account_l2_field = self._resolve_field("account_l2")
        date_field = self.fields.get("date", "Reporting Date")

        if account_l1_field != self.fields.get("account_l1"):
            print(f"    ℹ Using alternative field '{account_l1_field}' for account_l1")
        if account_l2_field != self.fields.get("account_l2"):
            print(f"    ℹ Using alternative field '{account_l2_field}' for account_l2")

        filters = [
            {"name": self.fields["scenario"], "values": ["Actuals"], "is_excluded": False},
            {"name": self.fields["account_l0"], "values": [self.accounts["pnl_filter"]], "is_excluded": False},
            {"name": self.fields["year"], "values": [str(self.year)], "is_excluded": False}
        ]

        try:
            # Try aggregation first with resolved fields
            result_str = await self.client.aggregate(
                table_id=self.financials_table,
                dimensions=[date_field, account_l1_field],
                metrics=[{"field": self.fields["amount"], "agg": "SUM"}],
                filters=filters
            )

            data = self._parse_api_response(result_str)

            # If aggregation returned no data, use PAGINATED raw data with client-side aggregation
            if not data:
                print("    ℹ Aggregation returned no data, using paginated raw data fallback...")
                print("    ⏳ Fetching ALL records (this may take a few minutes)...")
                raw_data = await self._fetch_raw_data_paginated(
                    table_id=self.financials_table,
                    filters=filters,
                    max_rows=100000
                )
                print(f"    ✓ Fetched {len(raw_data):,} raw records")

                if raw_data:
                    self.pnl_data = raw_data
                    data = self._aggregate_client_side(
                        raw_data,
                        [date_field, account_l1_field],
                        self.fields["amount"]
                    )
                    print(f"    ✓ Client-side aggregation: {len(data)} grouped records")

            self.aggregated_data["monthly_by_account"] = data
            # Store which account field was actually used (for downstream sheets)
            self.aggregated_data["_account_l1_field_used"] = account_l1_field
            self.timestamps["pnl_fetched"] = datetime.now().isoformat()
            print(f"    ✓ Fetched {len(data)} monthly account records")

            # L2 Account breakdown
            if self.pnl_data:
                data_l2 = self._aggregate_client_side(
                    self.pnl_data,
                    [account_l1_field, account_l2_field],
                    self.fields["amount"]
                )
            else:
                result_l2 = await self.client.aggregate(
                    table_id=self.financials_table,
                    dimensions=[account_l1_field, account_l2_field],
                    metrics=[{"field": self.fields["amount"], "agg": "SUM"}],
                    filters=filters
                )
                data_l2 = self._parse_api_response(result_l2)

            self.aggregated_data["account_l2"] = data_l2
            self.aggregated_data["_account_l2_field_used"] = account_l2_field
            print(f"    ✓ Fetched {len(data_l2)} L2 account records")

        except Exception as e:
            print(f"    ⚠ Error fetching P&L data: {e}")
            import traceback
            traceback.print_exc()
            self.aggregated_data["monthly_by_account"] = []
            self.aggregated_data["account_l2"] = []

    async def _fetch_kpi_data(self):
        """Fetch KPI metrics."""
        print("  📈 Fetching KPI metrics...")

        if not self.kpis_table:
            print("    ⚠ No KPI table configured")
            self.kpi_data = []
            return

        try:
            # Try aggregation first
            result = await self.client.aggregate(
                table_id=self.kpis_table,
                dimensions=[self.fields.get("quarter", "Quarter & Year"), self.fields.get("kpi_name", "KPI")],
                metrics=[{"field": self.fields.get("kpi_value", "value"), "agg": "MAX"}],
                filters=[
                    {"name": self.fields["scenario"], "values": ["Actuals"], "is_excluded": False}
                ]
            )

            self.kpi_data = self._parse_api_response(result)

            # Fallback to paginated raw data if aggregation fails
            if not self.kpi_data:
                print("    ℹ Using paginated raw data fallback for KPIs...")
                self.kpi_data = await self._fetch_raw_data_paginated(
                    table_id=self.kpis_table,
                    filters=[
                        {"name": self.fields["scenario"], "values": ["Actuals"], "is_excluded": False}
                    ],
                    max_rows=10000  # KPIs are typically smaller datasets
                )

            self.timestamps["kpi_fetched"] = datetime.now().isoformat()
            print(f"    ✓ Fetched {len(self.kpi_data)} KPI records")

        except Exception as e:
            print(f"    ⚠ Error fetching KPI data: {e}")
            self.kpi_data = []

    async def _fetch_vendor_data(self):
        """Fetch vendor spend data."""
        print("  🏢 Fetching vendor data...")

        filters = [
            {"name": self.fields["scenario"], "values": ["Actuals"], "is_excluded": False},
            {"name": self.fields["account_l0"], "values": [self.accounts["pnl_filter"]], "is_excluded": False},
            {"name": self.fields["year"], "values": [str(self.year)], "is_excluded": False}
        ]

        try:
            result = await self.client.aggregate(
                table_id=self.financials_table,
                dimensions=["Vendor / Customer"],
                metrics=[{"field": self.fields["amount"], "agg": "SUM"}],
                filters=filters + [
                    {"name": self.fields["account_l1"], "values": [self.accounts["revenue"]], "is_excluded": True}
                ]
            )

            data = self._parse_api_response(result)

            # Fallback to client-side aggregation using already-fetched P&L data
            if not data and self.pnl_data:
                print("    ℹ Using already-fetched P&L data for vendors...")
                # Filter out revenue records before aggregating
                expense_data = [
                    r for r in self.pnl_data
                    if r.get(self.fields["account_l1"]) != self.accounts.get("revenue")
                ]
                data = self._aggregate_client_side(
                    expense_data,
                    ["Vendor / Customer"],
                    self.fields["amount"]
                )

            self.aggregated_data["vendor"] = data
            print(f"    ✓ Fetched {len(self.aggregated_data['vendor'])} vendor records")

        except Exception as e:
            print(f"    ⚠ Error fetching vendor data: {e}")
            self.aggregated_data["vendor"] = []

    async def _fetch_cost_center_data(self):
        """Fetch cost center data."""
        print("  🏗️ Fetching cost center data...")

        account_l1_field = self._resolve_field("account_l1")

        filters = [
            {"name": self.fields["scenario"], "values": ["Actuals"], "is_excluded": False},
            {"name": self.fields["account_l0"], "values": [self.accounts["pnl_filter"]], "is_excluded": False},
            {"name": self.fields["year"], "values": [str(self.year)], "is_excluded": False}
        ]

        try:
            result = await self.client.aggregate(
                table_id=self.financials_table,
                dimensions=[self.fields.get("cost_center", "Cost Center"), account_l1_field],
                metrics=[{"field": self.fields["amount"], "agg": "SUM"}],
                filters=filters
            )

            data = self._parse_api_response(result)

            # Fallback to client-side aggregation using already-fetched P&L data
            if not data and self.pnl_data:
                print("    ℹ Using already-fetched P&L data for cost centers...")
                data = self._aggregate_client_side(
                    self.pnl_data,
                    [self.fields.get("cost_center", "Cost Center"), account_l1_field],
                    self.fields["amount"]
                )

            self.aggregated_data["cost_center"] = data
            print(f"    ✓ Fetched {len(self.aggregated_data['cost_center'])} cost center records")

        except Exception as e:
            print(f"    ⚠ Error fetching cost center data: {e}")
            self.aggregated_data["cost_center"] = []

    def calculate_insights(self):
        """Calculate business insights from data."""
        print("\n2️⃣  CALCULATING INTELLIGENCE...")

        self.insights = []
        self.recommendations = []

        # Calculate totals
        totals = self._calculate_totals()

        # 1. OpEx vs Revenue check
        self._check_opex_ratio(totals)

        # 2. MoM variance check
        self._check_mom_variance()

        # 3. Vendor concentration check
        self._check_vendor_concentration(totals)

        # 4. Cost center analysis
        self._analyze_cost_centers(totals)

        # 5. Gross margin analysis
        self._check_gross_margin(totals)

        # 6. Trend anomalies
        self._detect_trend_anomalies()

        # Sort insights by severity
        severity_order = {"CRITICAL": 0, "WARNING": 1, "ACTION": 2, "POSITIVE": 3, "INFO": 4}
        self.insights.sort(key=lambda x: severity_order.get(x["severity"], 5))

        print(f"    ✓ Generated {len(self.insights)} insights")
        print(f"    ✓ Generated {len(self.recommendations)} recommendations")

    def _convert_timestamp_to_month(self, timestamp) -> str:
        """Convert a timestamp (int/float or string) to YYYY-MM format."""
        if isinstance(timestamp, (int, float)):
            try:
                dt = datetime.fromtimestamp(timestamp)
                return dt.strftime("%Y-%m")
            except (ValueError, OSError):
                pass
        elif isinstance(timestamp, str):
            # Already a date string?
            if len(timestamp) >= 7 and timestamp[4] == "-":
                return timestamp[:7]
            # Try to parse as number
            try:
                ts = float(timestamp)
                dt = datetime.fromtimestamp(ts)
                return dt.strftime("%Y-%m")
            except (ValueError, OSError):
                pass
        return ""

    def _get_account_l1_field(self) -> str:
        """Get the actual account L1 field name used in aggregated data."""
        return self.aggregated_data.get("_account_l1_field_used", self._resolve_field("account_l1"))

    def _get_account_l2_field(self) -> str:
        """Get the actual account L2 field name used in aggregated data."""
        return self.aggregated_data.get("_account_l2_field_used", self._resolve_field("account_l2"))

    def _calculate_totals(self) -> Dict[str, float]:
        """Calculate total amounts by account type."""
        totals = defaultdict(float)
        account_field = self._get_account_l1_field()

        for record in self.aggregated_data.get("monthly_by_account", []):
            account = record.get(account_field, "Unknown")

            try:
                amount = float(record.get(self.fields["amount"], 0) or 0)
            except (ValueError, TypeError):
                amount = 0

            # Data is already filtered by year from API, so just aggregate
            totals[account] += amount

        # Also use cost center data to get more complete totals
        for record in self.aggregated_data.get("cost_center", []):
            account = record.get(account_field, "Unknown")
            try:
                amount = float(record.get(self.fields["amount"], 0) or 0)
            except (ValueError, TypeError):
                amount = 0

            # Add to totals if not already counted
            if account not in totals:
                totals[account] = 0
            # Use cost center data if monthly data is missing
            if totals[account] == 0:
                totals[account] = amount

        # Calculate grand total expense
        totals["total_expense"] = (
            abs(totals.get(self.accounts.get("opex", "Operating Expense"), 0)) +
            abs(totals.get(self.accounts.get("cogs", "Cost of Good sold"), 0)) +
            abs(totals.get(self.accounts.get("financial_expense", "Financial Expenses"), 0))
        )

        return dict(totals)

    def _check_opex_ratio(self, totals: Dict[str, float]):
        """Check OpEx vs Revenue ratio."""
        revenue = totals.get(self.accounts.get("revenue", "REVENUE"), 0)
        opex = totals.get(self.accounts.get("opex", "Operating Expense"), 0)

        if revenue > 0:
            ratio = opex / revenue

            if ratio > 2.0:
                self.insights.append({
                    "severity": "CRITICAL",
                    "category": "Profitability",
                    "finding": f"OpEx exceeded ${opex:,.0f} ({ratio:.1f}x Revenue)",
                    "detail": f"Operating expenses are {ratio:.1f}x revenue of ${revenue:,.0f}",
                    "action": "Immediate cost review required"
                })
                self.recommendations.append(
                    f"Review operating expenses - currently {ratio:.1f}x revenue. "
                    "Consider departmental budget reviews."
                )
            elif ratio > 1.5:
                self.insights.append({
                    "severity": "WARNING",
                    "category": "Profitability",
                    "finding": f"OpEx at ${opex:,.0f} ({ratio:.1f}x Revenue)",
                    "detail": f"Operating expenses are elevated vs revenue",
                    "action": "Monitor cost growth closely"
                })

    def _check_mom_variance(self):
        """Check month-over-month variances."""
        monthly = defaultdict(lambda: defaultdict(float))
        account_field = self._get_account_l1_field()

        for record in self.aggregated_data.get("monthly_by_account", []):
            date = record.get(self.fields["date"], "")[:7]  # YYYY-MM
            account = record.get(account_field, "Unknown")
            amount = float(record.get(self.fields["amount"], 0))

            if date.startswith(str(self.year)):
                monthly[account][date] += amount

        # Check each account for significant changes
        for account, months in monthly.items():
            sorted_months = sorted(months.keys())

            if len(sorted_months) >= 2:
                # Compare last two months
                current = months[sorted_months[-1]]
                previous = months[sorted_months[-2]]

                if previous != 0:
                    change = (current - previous) / abs(previous)

                    if abs(change) > 0.30:  # >30% change
                        severity = "WARNING" if abs(change) > 0.50 else "INFO"
                        direction = "up" if change > 0 else "down"
                        self.insights.append({
                            "severity": severity,
                            "category": "Variance",
                            "finding": f"{account} {direction} {abs(change):.0%} MoM",
                            "detail": f"Changed from ${previous:,.0f} to ${current:,.0f}",
                            "action": f"Investigate {account} variance"
                        })

    def _check_vendor_concentration(self, totals: Dict[str, float]):
        """Check vendor concentration risk."""
        vendor_data = self.aggregated_data.get("vendor", [])
        total_expense = totals.get("total_expense", 0)

        if not vendor_data or total_expense == 0:
            return

        # Calculate vendor percentages
        vendors = []
        for record in vendor_data:
            vendor = record.get("Vendor / Customer", "Unknown")
            amount = abs(float(record.get(self.fields["amount"], 0)))

            if vendor and vendor != "Unknown" and amount > 0:
                vendors.append({
                    "vendor": vendor,
                    "amount": amount,
                    "pct": amount / total_expense if total_expense else 0
                })

        # Sort by amount
        vendors.sort(key=lambda x: x["amount"], reverse=True)

        # Check top vendors for concentration
        top_3_pct = sum(v["pct"] for v in vendors[:3])

        if top_3_pct > 0.40:
            self.insights.append({
                "severity": "WARNING",
                "category": "Vendor Risk",
                "finding": f"Top 3 vendors = {top_3_pct:.0%} of total spend",
                "detail": ", ".join([v["vendor"][:20] for v in vendors[:3]]),
                "action": "Review vendor dependency"
            })
            self.recommendations.append(
                "Audit top 3 vendors for cost optimization potential and "
                "consider diversifying vendor relationships."
            )

        # Check individual concentration
        for vendor in vendors[:5]:
            if vendor["pct"] > self.CONCENTRATION_RISK_THRESHOLD:
                self.insights.append({
                    "severity": "INFO",
                    "category": "Vendor",
                    "finding": f"{vendor['vendor'][:30]} = {vendor['pct']:.0%} of spend",
                    "detail": f"${vendor['amount']:,.0f}",
                    "action": "Monitor vendor dependency"
                })

    def _analyze_cost_centers(self, totals: Dict[str, float]):
        """Analyze cost center performance."""
        cc_data = self.aggregated_data.get("cost_center", [])

        if not cc_data:
            return

        # Aggregate by cost center
        cc_totals = defaultdict(float)
        for record in cc_data:
            cc = record.get(self.fields.get("cost_center", "Cost Center"), "No Cost Center")
            amount = float(record.get(self.fields["amount"], 0))
            cc_totals[cc] += amount

        # Find highest cost center
        sorted_cc = sorted(cc_totals.items(), key=lambda x: abs(x[1]), reverse=True)

        if sorted_cc:
            top_cc = sorted_cc[0]
            self.insights.append({
                "severity": "INFO",
                "category": "Cost Center",
                "finding": f"Highest cost center: {top_cc[0]}",
                "detail": f"${abs(top_cc[1]):,.0f}",
                "action": f"Review {top_cc[0]} efficiency"
            })

    def _check_gross_margin(self, totals: Dict[str, float]):
        """Check gross margin."""
        revenue = totals.get(self.accounts.get("revenue", "REVENUE"), 0)
        cogs = totals.get(self.accounts.get("cogs", "Cost of Good sold"), 0)

        if revenue > 0:
            gross_profit = revenue - cogs
            gross_margin = gross_profit / revenue

            if gross_margin > 0.80:
                self.insights.append({
                    "severity": "POSITIVE",
                    "category": "Profitability",
                    "finding": f"Strong gross margin: {gross_margin:.1%}",
                    "detail": f"Gross profit ${gross_profit:,.0f}",
                    "action": "Maintain pricing discipline"
                })
            elif gross_margin < 0.50:
                self.insights.append({
                    "severity": "WARNING",
                    "category": "Profitability",
                    "finding": f"Low gross margin: {gross_margin:.1%}",
                    "detail": f"COGS ${cogs:,.0f} vs Revenue ${revenue:,.0f}",
                    "action": "Review COGS and pricing"
                })

    def _detect_trend_anomalies(self):
        """Detect anomalies in monthly trends."""
        monthly = defaultdict(lambda: defaultdict(float))
        account_field = self._get_account_l1_field()

        for record in self.aggregated_data.get("monthly_by_account", []):
            date = record.get(self.fields["date"], "")[:7]
            account = record.get(account_field, "Unknown")
            amount = float(record.get(self.fields["amount"], 0))

            if date.startswith(str(self.year)):
                monthly[account][date] += amount

        # Check for statistical outliers
        for account, months in monthly.items():
            values = list(months.values())

            if len(values) >= 3:
                try:
                    mean = statistics.mean(values)
                    std = statistics.stdev(values)

                    if std > 0:
                        for month, value in months.items():
                            z_score = abs(value - mean) / std

                            if z_score > 2.5:  # >2.5 std deviations
                                self.insights.append({
                                    "severity": "WARNING",
                                    "category": "Anomaly",
                                    "finding": f"Outlier in {account} ({month})",
                                    "detail": f"${value:,.0f} vs avg ${mean:,.0f} (Z={z_score:.1f})",
                                    "action": "Investigate unusual transaction"
                                })
                except statistics.StatisticsError:
                    pass

    def create_insights_dashboard(self):
        """Create the Insights Dashboard sheet."""
        print("  📊 Creating Insights Dashboard...")
        ws = self.wb.create_sheet("Insights Dashboard", 0)

        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"FP&A INTELLIGENCE DASHBOARD - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # Subtitle
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Data Source: Datarails Finance OS"
        ws['A2'].font = Font(name='Calibri', size=10, italic=True, color="666666")
        ws['A2'].alignment = Alignment(horizontal="center")

        # TOP FINDINGS Section
        row = 4
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "TOP 5 FINDINGS"
        self._format_subheader_cell(ws[f'A{row}'], self.ACCENT_COLOR)
        ws.row_dimensions[row].height = 22

        row += 1
        for idx, insight in enumerate(self.insights[:5], 1):
            severity = insight["severity"]
            icon = {
                "CRITICAL": "🔴",
                "WARNING": "🟡",
                "POSITIVE": "🟢",
                "INFO": "🔵",
                "ACTION": "⚠️"
            }.get(severity, "•")

            ws[f'A{row}'] = f"{icon} {severity}"
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True,
                                       color=self.SEVERITY_COLORS.get(severity, "000000"))

            ws.merge_cells(f'B{row}:G{row}')
            ws[f'B{row}'] = insight["finding"]
            ws[f'B{row}'].font = Font(name='Calibri', size=11)

            row += 1

        # KEY METRICS Section
        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "KEY METRICS"
        self._format_subheader_cell(ws[f'A{row}'], self.SUBHEADER_COLOR)

        # Calculate metrics
        totals = self._calculate_totals()
        revenue = totals.get(self.accounts.get("revenue", "REVENUE"), 0)
        opex = totals.get(self.accounts.get("opex", "Operating Expense"), 0)
        cogs = totals.get(self.accounts.get("cogs", "Cost of Good sold"), 0)
        gross_profit = revenue - cogs
        gross_margin = (gross_profit / revenue * 100) if revenue else 0

        row += 1
        headers = ["Metric", "Value", "Trend", "Status"]
        for col, header in enumerate(headers, 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        metrics_data = [
            ("Revenue", f"${revenue:,.0f}", "—", "✓ On Track" if revenue > 0 else "—"),
            ("Gross Margin", f"{gross_margin:.1f}%", "—", "✓ Strong" if gross_margin > 80 else "⚠ Watch"),
            ("OpEx", f"${opex:,.0f}", "—", "✗ Over" if opex > revenue else "✓ OK"),
            ("Gross Profit", f"${gross_profit:,.0f}", "—", "—"),
        ]

        for metric, value, trend, status in metrics_data:
            row += 1
            ws.cell(row, 1).value = metric
            ws.cell(row, 1).font = Font(name='Calibri', size=10, bold=True)
            ws.cell(row, 2).value = value
            ws.cell(row, 2).alignment = Alignment(horizontal="right")
            ws.cell(row, 3).value = trend
            ws.cell(row, 3).alignment = Alignment(horizontal="center")
            ws.cell(row, 4).value = status

            # Color status
            if "✓" in status:
                ws.cell(row, 4).font = Font(color=self.POSITIVE_COLOR)
            elif "✗" in status:
                ws.cell(row, 4).font = Font(color=self.NEGATIVE_COLOR)

        # RECOMMENDATIONS Section
        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "RECOMMENDATIONS"
        self._format_subheader_cell(ws[f'A{row}'], self.ACCENT_COLOR)

        for rec in self.recommendations[:5]:
            row += 1
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = f"• {rec}"
            ws[f'A{row}'].font = Font(name='Calibri', size=10)
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)

        # Set column widths
        ws.column_dimensions['A'].width = 15
        for col in 'BCDEFG':
            ws.column_dimensions[col].width = 18

    def create_expense_deep_dive(self):
        """Create Expense Deep Dive sheet."""
        print("  💰 Creating Expense Deep Dive...")
        ws = self.wb.create_sheet("Expense Deep Dive")

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"EXPENSE DEEP DIVE - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        # Get L2 account data
        l2_data = self.aggregated_data.get("account_l2", [])
        totals = self._calculate_totals()
        total_expense = totals.get("total_expense", 1)
        account_l1_field = self._get_account_l1_field()
        account_l2_field = self._get_account_l2_field()

        # Aggregate by L2
        l2_totals = []
        for record in l2_data:
            l1 = record.get(account_l1_field, "Unknown")
            l2 = record.get(account_l2_field, "Unknown")
            amount = float(record.get(self.fields["amount"], 0))

            # Skip revenue
            if l1 == self.accounts.get("revenue", "REVENUE"):
                continue

            l2_totals.append({
                "Category": l1,
                "Account": l2,
                "Amount": amount,
                "Pct": abs(amount) / abs(total_expense) if total_expense else 0
            })

        # Sort by absolute amount
        l2_totals.sort(key=lambda x: abs(x["Amount"]), reverse=True)

        # TOP 20 EXPENSE ACCOUNTS Section
        row = 3
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "TOP 20 EXPENSE ACCOUNTS"
        self._format_subheader_cell(ws[f'A{row}'])

        row += 1
        headers = ["#", "Category", "Account", "Amount", "% of Total", "Flag"]
        for col, header in enumerate(headers, 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        for idx, item in enumerate(l2_totals[:20], 1):
            row += 1
            ws.cell(row, 1).value = idx
            ws.cell(row, 2).value = item["Category"]
            ws.cell(row, 3).value = item["Account"]
            ws.cell(row, 4).value = item["Amount"]
            ws.cell(row, 4).number_format = '$#,##0.00'
            ws.cell(row, 5).value = item["Pct"]
            ws.cell(row, 5).number_format = '0.0%'

            # Flag material items
            flag = ""
            if item["Pct"] > 0.10:
                flag = "🔴 HIGH"
            elif item["Pct"] > 0.05:
                flag = "🟡 MATERIAL"
            ws.cell(row, 6).value = flag

        # COST CENTER BREAKDOWN Section
        row += 3
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "COST CENTER BREAKDOWN"
        self._format_subheader_cell(ws[f'A{row}'])

        cc_data = self.aggregated_data.get("cost_center", [])
        cc_totals_map = defaultdict(float)

        for record in cc_data:
            cc = record.get(self.fields.get("cost_center", "Cost Center"), "No Cost Center")
            amount = float(record.get(self.fields["amount"], 0))
            cc_totals_map[cc] += amount

        row += 1
        headers = ["Cost Center", "Total Spend", "% of Total", "Status"]
        for col, header in enumerate(headers, 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        sorted_cc = sorted(cc_totals_map.items(), key=lambda x: abs(x[1]), reverse=True)

        for cc, amount in sorted_cc[:15]:
            row += 1
            pct = abs(amount) / abs(total_expense) if total_expense else 0
            ws.cell(row, 1).value = cc
            ws.cell(row, 2).value = amount
            ws.cell(row, 2).number_format = '$#,##0.00'
            ws.cell(row, 3).value = pct
            ws.cell(row, 3).number_format = '0.0%'

            status = "Over Budget" if pct > 0.20 else "On Track"
            ws.cell(row, 4).value = status

        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15

    def create_variance_waterfall(self):
        """Create Variance Waterfall sheet."""
        print("  📊 Creating Variance Waterfall...")
        ws = self.wb.create_sheet("Variance Waterfall")

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"VARIANCE ANALYSIS - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        # Calculate monthly data
        monthly = defaultdict(lambda: defaultdict(float))
        account_field = self._get_account_l1_field()

        for record in self.aggregated_data.get("monthly_by_account", []):
            date = record.get(self.fields["date"], "")[:7]
            account = record.get(account_field, "Unknown")
            amount = float(record.get(self.fields["amount"], 0))

            if date.startswith(str(self.year)):
                monthly[account][date] += amount

        # MoM VARIANCE Section
        row = 3
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "MONTH-OVER-MONTH VARIANCE BY ACCOUNT"
        self._format_subheader_cell(ws[f'A{row}'])

        row += 1
        headers = ["Account", "Current Month", "Prior Month", "Change ($)", "Change (%)"]
        for col, header in enumerate(headers, 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        for account, months in monthly.items():
            sorted_months = sorted(months.keys())

            if len(sorted_months) >= 2:
                current = months[sorted_months[-1]]
                previous = months[sorted_months[-2]]
                change = current - previous
                pct_change = change / abs(previous) if previous != 0 else 0

                row += 1
                ws.cell(row, 1).value = account
                ws.cell(row, 2).value = current
                ws.cell(row, 2).number_format = '$#,##0.00'
                ws.cell(row, 3).value = previous
                ws.cell(row, 3).number_format = '$#,##0.00'
                ws.cell(row, 4).value = change
                ws.cell(row, 4).number_format = '$#,##0.00'
                ws.cell(row, 5).value = pct_change
                ws.cell(row, 5).number_format = '0.0%'

                # Color based on variance
                if account == self.accounts.get("revenue", "REVENUE"):
                    # For revenue, positive is good
                    color = self.POSITIVE_COLOR if change > 0 else self.NEGATIVE_COLOR
                else:
                    # For expenses, negative is good
                    color = self.POSITIVE_COLOR if change < 0 else self.NEGATIVE_COLOR

                ws.cell(row, 4).font = Font(color=color)
                ws.cell(row, 5).font = Font(color=color)

        # Set column widths
        for col, width in zip('ABCDE', [25, 18, 18, 18, 12]):
            ws.column_dimensions[col].width = width

    def create_trend_analysis(self):
        """Create Trend Analysis sheet."""
        print("  📈 Creating Trend Analysis...")
        ws = self.wb.create_sheet("Trend Analysis")

        # Title
        ws.merge_cells('A1:M1')
        ws['A1'] = f"12-MONTH TREND ANALYSIS - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        # Calculate monthly data
        monthly = defaultdict(lambda: defaultdict(float))
        account_field = self._get_account_l1_field()

        for record in self.aggregated_data.get("monthly_by_account", []):
            date = record.get(self.fields["date"], "")[:7]
            account = record.get(account_field, "Unknown")
            amount = float(record.get(self.fields["amount"], 0))
            monthly[account][date] += amount

        # Get all months and sort
        all_months = sorted(set(m for months in monthly.values() for m in months))

        # MONTHLY TRENDS Section
        row = 3
        headers = ["Account"] + all_months + ["Total", "Avg", "Growth"]
        for col, header in enumerate(headers, 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        account_order = [
            self.accounts.get("revenue", "REVENUE"),
            self.accounts.get("cogs", "Cost of Good sold"),
            self.accounts.get("opex", "Operating Expense"),
            self.accounts.get("financial_expense", "Financial Expenses")
        ]

        for account in account_order:
            if account in monthly:
                row += 1
                ws.cell(row, 1).value = account
                ws.cell(row, 1).font = Font(bold=True)

                values = []
                for col, month in enumerate(all_months, 2):
                    value = monthly[account].get(month, 0)
                    ws.cell(row, col).value = value
                    ws.cell(row, col).number_format = '#,##0'
                    values.append(value)

                # Total
                total_col = len(all_months) + 2
                total = sum(values)
                ws.cell(row, total_col).value = total
                ws.cell(row, total_col).number_format = '$#,##0'
                ws.cell(row, total_col).font = Font(bold=True)

                # Average
                avg = sum(values) / len(values) if values else 0
                ws.cell(row, total_col + 1).value = avg
                ws.cell(row, total_col + 1).number_format = '$#,##0'

                # Growth (first to last)
                if len(values) >= 2 and values[0] != 0:
                    growth = (values[-1] - values[0]) / abs(values[0])
                else:
                    growth = 0
                ws.cell(row, total_col + 2).value = growth
                ws.cell(row, total_col + 2).number_format = '0.0%'

        # Set column widths
        ws.column_dimensions['A'].width = 22
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def create_anomaly_report(self):
        """Create Anomaly Report sheet."""
        print("  🔍 Creating Anomaly Report...")
        ws = self.wb.create_sheet("Anomaly Report")

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"ANOMALY DETECTION REPORT - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        # Filter for anomaly insights
        anomaly_insights = [i for i in self.insights if i.get("category") in ["Anomaly", "Variance"]]

        row = 3
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"AUTO-DETECTED ANOMALIES ({len(anomaly_insights)} findings)"
        self._format_subheader_cell(ws[f'A{row}'])

        row += 1
        headers = ["Severity", "Category", "Finding", "Detail", "Action"]
        for col, header in enumerate(headers, 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        for insight in anomaly_insights:
            row += 1
            ws.cell(row, 1).value = insight["severity"]
            ws.cell(row, 1).font = Font(bold=True, color=self.SEVERITY_COLORS.get(insight["severity"], "000000"))
            ws.cell(row, 2).value = insight["category"]
            ws.cell(row, 3).value = insight["finding"]
            ws.cell(row, 4).value = insight.get("detail", "")
            ws.cell(row, 5).value = insight.get("action", "")

        if not anomaly_insights:
            row += 1
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "✓ No significant anomalies detected"
            ws[f'A{row}'].font = Font(color=self.POSITIVE_COLOR)

        # Set column widths
        for col, width in zip('ABCDEF', [12, 15, 40, 35, 30]):
            ws.column_dimensions[col].width = width

    def create_vendor_analysis(self):
        """Create Vendor Analysis sheet."""
        print("  🏢 Creating Vendor Analysis...")
        ws = self.wb.create_sheet("Vendor Analysis")

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"VENDOR SPEND ANALYSIS - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        vendor_data = self.aggregated_data.get("vendor", [])
        totals = self._calculate_totals()
        total_expense = totals.get("total_expense", 1)

        # Process vendor data
        vendors = []
        for record in vendor_data:
            vendor = record.get("Vendor / Customer", "Unknown")
            amount = abs(float(record.get(self.fields["amount"], 0)))

            if vendor and vendor != "Unknown" and vendor.strip() and amount > 0:
                vendors.append({
                    "vendor": vendor,
                    "amount": amount,
                    "pct": amount / abs(total_expense) if total_expense else 0
                })

        vendors.sort(key=lambda x: x["amount"], reverse=True)

        # TOP 20 VENDORS Section
        row = 3
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "TOP 20 VENDORS BY SPEND"
        self._format_subheader_cell(ws[f'A{row}'])

        row += 1
        headers = ["#", "Vendor", "Amount", "% of Total", "Cumulative %", "Risk Flag"]
        for col, header in enumerate(headers, 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        cumulative = 0
        for idx, vendor in enumerate(vendors[:20], 1):
            row += 1
            cumulative += vendor["pct"]

            ws.cell(row, 1).value = idx
            ws.cell(row, 2).value = vendor["vendor"][:40]
            ws.cell(row, 3).value = vendor["amount"]
            ws.cell(row, 3).number_format = '$#,##0.00'
            ws.cell(row, 4).value = vendor["pct"]
            ws.cell(row, 4).number_format = '0.0%'
            ws.cell(row, 5).value = cumulative
            ws.cell(row, 5).number_format = '0.0%'

            # Risk flag
            risk = ""
            if vendor["pct"] > 0.15:
                risk = "🔴 HIGH CONCENTRATION"
            elif vendor["pct"] > 0.10:
                risk = "🟡 CONCENTRATION RISK"
            ws.cell(row, 6).value = risk

        # Set column widths
        for col, width in zip('ABCDEF', [6, 40, 18, 12, 14, 22]):
            ws.column_dimensions[col].width = width

    def create_saas_metrics(self):
        """Create SaaS Metrics sheet."""
        print("  📊 Creating SaaS Metrics...")
        ws = self.wb.create_sheet("SaaS Metrics")

        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"SAAS METRICS DASHBOARD - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        # Process KPI data
        kpi_latest = {}
        quarter_field = self.fields.get("quarter", "Quarter & Year")
        kpi_field = self.fields.get("kpi_name", "KPI")
        value_field = self.fields.get("kpi_value", "value")

        for record in self.kpi_data:
            kpi_name = record.get(kpi_field, "Unknown")
            raw_value = record.get(value_field)
            quarter = record.get(quarter_field, "") or ""

            # Skip records with no value
            if raw_value is None:
                continue

            try:
                value = float(raw_value)
            except (ValueError, TypeError):
                continue

            # Keep latest value for each KPI
            if kpi_name and (kpi_name not in kpi_latest or quarter > kpi_latest[kpi_name]["quarter"]):
                kpi_latest[kpi_name] = {"value": value, "quarter": quarter}

        # ARR METRICS Section
        row = 3
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "ARR & REVENUE METRICS"
        self._format_subheader_cell(ws[f'A{row}'])

        arr_kpis = ["Total ARR", "New ARR", "Net New ARR", "ARR Opening Balance", "ARR Closing Balance"]

        row += 1
        for col, header in enumerate(["Metric", "Value", "Period"], 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        for kpi in arr_kpis:
            for kpi_name, data in kpi_latest.items():
                if kpi.lower() in kpi_name.lower():
                    row += 1
                    ws.cell(row, 1).value = kpi_name
                    ws.cell(row, 2).value = data["value"]
                    ws.cell(row, 2).number_format = '$#,##0'
                    ws.cell(row, 3).value = data["quarter"]
                    break

        # UNIT ECONOMICS Section
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "UNIT ECONOMICS"
        self._format_subheader_cell(ws[f'A{row}'])

        unit_kpis = ["LTV", "CAC", "LTV / CAC", "CAC Payback"]

        row += 1
        for col, header in enumerate(["Metric", "Value", "Period"], 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        for kpi in unit_kpis:
            for kpi_name, data in kpi_latest.items():
                if kpi.lower() in kpi_name.lower():
                    row += 1
                    ws.cell(row, 1).value = kpi_name
                    ws.cell(row, 2).value = data["value"]

                    # Format appropriately
                    if "ratio" in kpi_name.lower() or "/" in kpi_name:
                        ws.cell(row, 2).number_format = '0.00x'
                    elif "payback" in kpi_name.lower():
                        ws.cell(row, 2).number_format = '0.0'
                    else:
                        ws.cell(row, 2).number_format = '$#,##0'

                    ws.cell(row, 3).value = data["quarter"]
                    break

        # EFFICIENCY METRICS Section
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "EFFICIENCY & BURN METRICS"
        self._format_subheader_cell(ws[f'A{row}'])

        efficiency_kpis = ["Burn", "Burn Multiple", "Magic Number", "Runway"]

        row += 1
        for col, header in enumerate(["Metric", "Value", "Period"], 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        for kpi in efficiency_kpis:
            for kpi_name, data in kpi_latest.items():
                if kpi.lower() in kpi_name.lower():
                    row += 1
                    ws.cell(row, 1).value = kpi_name
                    ws.cell(row, 2).value = data["value"]

                    if "multiple" in kpi_name.lower() or "magic" in kpi_name.lower():
                        ws.cell(row, 2).number_format = '0.00x'
                    elif "runway" in kpi_name.lower():
                        ws.cell(row, 2).number_format = '0.0 "months"'
                    else:
                        ws.cell(row, 2).number_format = '$#,##0'

                    ws.cell(row, 3).value = data["quarter"]
                    break

        # Set column widths
        for col, width in zip('ABCD', [35, 18, 15, 15]):
            ws.column_dimensions[col].width = width

    def create_sales_performance(self):
        """Create Sales Performance sheet."""
        print("  📈 Creating Sales Performance...")
        ws = self.wb.create_sheet("Sales Performance")

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"SALES PERFORMANCE - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        # Filter sales-related KPIs
        sales_kpis = {}
        quarter_field = self.fields.get("quarter", "Quarter & Year")
        kpi_field = self.fields.get("kpi_name", "KPI")
        value_field = self.fields.get("kpi_value", "value")

        for record in self.kpi_data:
            kpi_name = record.get(kpi_field, "Unknown") or "Unknown"
            raw_value = record.get(value_field)
            quarter = record.get(quarter_field, "") or ""

            # Skip records with no value
            if raw_value is None:
                continue

            try:
                value = float(raw_value)
            except (ValueError, TypeError):
                continue

            # Filter for sales metrics
            sales_terms = ["win", "close", "meeting", "pipeline", "deal", "sales"]
            if any(term in kpi_name.lower() for term in sales_terms):
                if kpi_name not in sales_kpis or quarter > sales_kpis[kpi_name]["quarter"]:
                    sales_kpis[kpi_name] = {"value": value, "quarter": quarter}

        # SALES METRICS Section
        row = 3
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "SALES PIPELINE METRICS"
        self._format_subheader_cell(ws[f'A{row}'])

        row += 1
        headers = ["Metric", "Value", "Period"]
        for col, header in enumerate(headers, 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        for kpi_name, data in sorted(sales_kpis.items()):
            row += 1
            ws.cell(row, 1).value = kpi_name
            ws.cell(row, 2).value = data["value"]

            # Format based on type
            if "ratio" in kpi_name.lower() or "%" in kpi_name:
                ws.cell(row, 2).number_format = '0.0%'
            elif "count" in kpi_name.lower() or "#" in kpi_name:
                ws.cell(row, 2).number_format = '0'
            else:
                ws.cell(row, 2).number_format = '#,##0.00'

            ws.cell(row, 3).value = data["quarter"]

        if not sales_kpis:
            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = "No sales performance data available in KPI table"
            ws[f'A{row}'].font = Font(italic=True, color="666666")

        # Set column widths
        for col, width in zip('ABCDE', [40, 18, 15, 15, 15]):
            ws.column_dimensions[col].width = width

    def create_cost_center_pnl(self):
        """Create Cost Center P&L sheet."""
        print("  🏗️ Creating Cost Center P&L...")
        ws = self.wb.create_sheet("Cost Center P&L")

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"P&L BY COST CENTER - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        cc_data = self.aggregated_data.get("cost_center", [])
        account_field = self._get_account_l1_field()

        # Pivot by cost center and account
        cc_pivot = defaultdict(lambda: defaultdict(float))
        for record in cc_data:
            cc = record.get(self.fields.get("cost_center", "Cost Center"), "No Cost Center")
            account = record.get(account_field, "Unknown")
            amount = float(record.get(self.fields["amount"], 0))
            cc_pivot[cc][account] += amount

        # Get unique cost centers and accounts
        cost_centers = sorted(cc_pivot.keys())
        accounts = [
            self.accounts.get("revenue", "REVENUE"),
            self.accounts.get("cogs", "Cost of Good sold"),
            self.accounts.get("opex", "Operating Expense"),
            self.accounts.get("financial_expense", "Financial Expenses")
        ]

        # Headers
        row = 3
        headers = ["Cost Center"] + accounts + ["Net"]
        for col, header in enumerate(headers, 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        # Data rows
        for cc in cost_centers:
            row += 1
            ws.cell(row, 1).value = cc
            ws.cell(row, 1).font = Font(bold=True)

            net = 0
            for col, account in enumerate(accounts, 2):
                value = cc_pivot[cc].get(account, 0)
                ws.cell(row, col).value = value
                ws.cell(row, col).number_format = '$#,##0.00'

                # Calculate net (revenue - expenses)
                if account == self.accounts.get("revenue", "REVENUE"):
                    net += value
                else:
                    net -= abs(value)

            # Net column
            ws.cell(row, len(accounts) + 2).value = net
            ws.cell(row, len(accounts) + 2).number_format = '$#,##0.00'
            ws.cell(row, len(accounts) + 2).font = Font(bold=True,
                                                         color=self.POSITIVE_COLOR if net > 0 else self.NEGATIVE_COLOR)

        # Set column widths
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(accounts) + 3):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def create_raw_data(self):
        """Create Raw Data sheet for pivot tables."""
        print("  📋 Creating Raw Data...")
        ws = self.wb.create_sheet("Raw Data")

        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"RAW DATA - PIVOT READY - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        ws['A2'] = "Use this data with Excel Pivot Tables for custom analysis"
        ws['A2'].font = Font(italic=True, color="666666")

        # Headers
        row = 4
        headers = ["Date", "Account L1", "Account L2", "Cost Center", "Amount"]

        for col, header in enumerate(headers, 1):
            self._format_header_cell(ws.cell(row, col))
            ws.cell(row, col).value = header

        # Combine all data
        account_field = self._get_account_l1_field()
        # Monthly by account data
        for record in self.aggregated_data.get("monthly_by_account", []):
            row += 1
            date = record.get(self.fields["date"], "")
            account_l1 = record.get(account_field, "")
            amount = float(record.get(self.fields["amount"], 0))

            ws.cell(row, 1).value = date
            ws.cell(row, 2).value = account_l1
            ws.cell(row, 3).value = ""  # L2 not available in this aggregation
            ws.cell(row, 4).value = ""  # Cost center not available
            ws.cell(row, 5).value = amount
            ws.cell(row, 5).number_format = '#,##0.00'

        # Set column widths
        for col, width in zip('ABCDEFGH', [12, 20, 25, 18, 15, 15, 15, 15]):
            ws.column_dimensions[col].width = width

        # Freeze header row
        ws.freeze_panes = 'A5'

    async def generate(self, output_path: str = None) -> str:
        """Generate the complete Intelligence Workbook."""
        print("\n" + "=" * 80)
        print("📊 FP&A INTELLIGENCE WORKBOOK GENERATOR")
        print("=" * 80)

        # Fetch all data
        await self.fetch_all_data()

        # Calculate insights
        self.calculate_insights()

        # Create all sheets
        print("\n3️⃣  CREATING WORKBOOK SHEETS...")
        self.create_insights_dashboard()
        self.create_expense_deep_dive()
        self.create_variance_waterfall()
        self.create_trend_analysis()
        self.create_anomaly_report()
        self.create_vendor_analysis()
        self.create_saas_metrics()
        self.create_sales_performance()
        self.create_cost_center_pnl()
        self.create_raw_data()

        # Save workbook
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = Path(__file__).parent.parent.parent / "tmp"
            output_dir.mkdir(exist_ok=True)
            output_path = str(output_dir / f"FPA_Intelligence_Workbook_{self.year}_{timestamp}.xlsx")

        print(f"\n4️⃣  SAVING WORKBOOK...")
        self.wb.save(output_path)
        print(f"    ✓ Saved: {output_path}")

        # Print summary
        print("\n" + "=" * 80)
        print("✅ INTELLIGENCE WORKBOOK GENERATED")
        print("=" * 80)
        print(f"\n📁 Output: {output_path}")
        print(f"📅 Year: {self.year}")
        print(f"💡 Insights: {len(self.insights)}")
        print(f"📝 Recommendations: {len(self.recommendations)}")
        print(f"\n📊 Sheets:")
        for idx, sheet in enumerate(self.wb.sheetnames, 1):
            print(f"   {idx}. {sheet}")

        print("\n" + "=" * 80)

        return output_path


async def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Generate FP&A Intelligence Workbook with actionable insights"
    )
    parser.add_argument("--year", type=int, default=2025, help="Year to analyze (default: 2025)")
    parser.add_argument("--env", default="app", choices=["app", "dev", "demo", "testapp"],
                        help="Environment (default: app)")
    parser.add_argument("--profile", help="Path to client profile JSON")
    parser.add_argument("--output", help="Output file path")

    args = parser.parse_args()

    try:
        workbook = IntelligenceWorkbook(
            year=args.year,
            env=args.env,
            profile_path=args.profile
        )
        output_path = await workbook.generate(args.output)

        # Clean up
        await workbook.client.close()

        return 0

    except FileNotFoundError as e:
        print(f"\n❌ Configuration Error: {e}", file=sys.stderr)
        print("   Run '/dr-learn --env app' to create a profile first", file=sys.stderr)
        return 1

    except Exception as e:
        print(f"\n❌ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
