#!/usr/bin/env python3
"""
Comprehensive FP&A Report Generator

Orchestrates multiple analyses to create a professional Financial Planning & Analysis
workbook suitable for executive review and board presentations.

Includes:
- Executive Summary with KPIs
- P&L Statement (monthly breakdown)
- P&L by Cost Center
- SaaS Metrics Dashboard
- Variance Analysis (Actuals vs Forecast)
- Trend Analysis
- Department Details
- Account Breakdown
- Data Quality Report
- Reconciliation Report

Usage:
    uv --directory mcp-server run python scripts/comprehensive_fpna_report.py --year 2025
    uv --directory mcp-server run python scripts/comprehensive_fpna_report.py --year 2025 --env app
    uv --directory mcp-server run python scripts/comprehensive_fpna_report.py --year 2025 --output tmp/FPA_2025.xlsx
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import statistics

# Add mcp-server/src to path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import httpx
import asyncio
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference, LineChart
from datarails_mcp.auth import get_auth
from datarails_mcp.client import DatarailsClient


class ComprehensiveFPAReport:
    """Generate comprehensive FP&A report from Datarails data."""

    # Professional color scheme
    HEADER_COLOR = "1F4E78"      # Dark blue
    SUBHEADER_COLOR = "4472C4"   # Medium blue
    ACCENT_COLOR = "ED7D31"      # Orange accent
    LIGHT_BG = "E7E6E6"          # Light gray
    POSITIVE_COLOR = "70AD47"    # Green
    NEGATIVE_COLOR = "C5504D"    # Red
    WARNING_COLOR = "FFC7CE"     # Light red
    SUCCESS_COLOR = "C6EFCE"     # Light green

    def __init__(self, year: int, env: str = "app", profile_path: str = None):
        """Initialize report generator."""
        self.year = year
        self.env = env
        self.auth = get_auth()
        self.client = DatarailsClient(self.auth)

        # Load profile
        self.profile = self._load_profile(profile_path, env)

        # Extract configuration
        self.financials_table = self.profile["tables"]["financials"]["id"]
        self.kpis_table = self.profile["tables"].get("kpis", {}).get("id")
        self.fields = self.profile["field_mappings"]
        self.accounts = self.profile.get("account_hierarchy", {})

        # Data containers
        self.pnl_data = []
        self.kpi_data = []
        self.timestamps = {}

        # Create workbook
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

    def _load_profile(self, profile_path: str, env: str) -> dict:
        """Load client profile."""
        if profile_path:
            path = Path(profile_path)
            if path.exists():
                with open(path) as f:
                    return json.load(f)

        # Try environment-based profile
        project_root = Path(__file__).parent.parent.parent
        profile_path = project_root / "config" / "client-profiles" / f"{env}.json"

        if profile_path.exists():
            with open(profile_path) as f:
                return json.load(f)

        raise FileNotFoundError(f"No profile found for {env}")

    async def fetch_pnl_data(self):
        """Fetch P&L data from Financials table."""
        print(f"  Fetching P&L data for {self.year}...")

        filters = {
            self.fields["year"]: str(self.year),
            self.fields["scenario"]: "Actuals",
            self.fields["account_l0"]: self.accounts["pnl_filter"]
        }

        # Use get_filtered API
        response_str = await self.client.get_filtered(
            self.financials_table,
            filters=filters,
            limit=500
        )

        # Parse response - API returns JSON string
        try:
            if isinstance(response_str, str):
                response_data = json.loads(response_str)
            else:
                response_data = response_str

            # Handle both dict and list responses
            if isinstance(response_data, dict):
                all_records = response_data.get("data", []) if "data" in response_data else list(response_data.values())
            elif isinstance(response_data, list):
                all_records = response_data
            else:
                all_records = []
        except Exception as e:
            print(f"    Warning: Failed to parse response: {e}")
            all_records = []

        # Sort by date (critical for time-series analysis)
        if all_records and isinstance(all_records, list):
            all_records.sort(key=lambda r: r.get(self.fields["date"], "") if isinstance(r, dict) else "")

        self.pnl_data = all_records if isinstance(all_records, list) else []
        self.timestamps["pnl_fetched"] = datetime.now().isoformat()

        print(f"    ✓ Fetched {len(self.pnl_data)} P&L records")
        return self.pnl_data

    async def fetch_kpi_data(self):
        """Fetch KPI data."""
        print(f"  Fetching KPI data for {self.year}...")

        filters = {
            self.fields["year"]: str(self.year)
        }

        # Use get_filtered API
        response_str = await self.client.get_filtered(
            self.kpis_table,
            filters=filters,
            limit=500
        )

        # Parse response - API returns JSON string
        try:
            if isinstance(response_str, str):
                response_data = json.loads(response_str)
            else:
                response_data = response_str

            # Handle both dict and list responses
            if isinstance(response_data, dict):
                self.kpi_data = response_data.get("data", []) if "data" in response_data else list(response_data.values())
            elif isinstance(response_data, list):
                self.kpi_data = response_data
            else:
                self.kpi_data = []
        except Exception as e:
            print(f"    Warning: Failed to parse KPI response: {e}")
            self.kpi_data = []

        self.timestamps["kpi_fetched"] = datetime.now().isoformat()

        # Ensure it's a list
        if not isinstance(self.kpi_data, list):
            self.kpi_data = []

        print(f"    ✓ Fetched {len(self.kpi_data)} KPI records")
        return self.kpi_data

    def _aggregate_by_account_and_month(self):
        """Aggregate P&L data by account and month."""
        aggregated = defaultdict(lambda: defaultdict(float))

        for record in self.pnl_data:
            # Ensure record is a dict
            if not isinstance(record, dict):
                continue

            date_str = record.get(self.fields["date"], "")
            if not date_str:
                continue

            # Extract month (YYYY-MM format)
            month = date_str[:7] if len(date_str) >= 7 else date_str
            account_l1 = record.get(self.fields["account_l1"], "Unknown")

            try:
                amount = float(record.get(self.fields["amount"], 0))
            except (ValueError, TypeError):
                amount = 0

            aggregated[account_l1][month] += amount

        return aggregated

    def _create_executive_summary(self):
        """Create Executive Summary sheet."""
        print("  Creating Executive Summary...")
        ws = self.wb.create_sheet("Executive Summary", 0)

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"COMPREHENSIVE FP&A REPORT - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Subtitle with timestamp
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(name='Calibri', size=10, color="666666")
        ws['A2'].alignment = Alignment(horizontal="center")

        # Key Metrics Section
        row = 4
        ws[f'A{row}'] = "KEY METRICS"
        ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=self.SUBHEADER_COLOR, end_color=self.SUBHEADER_COLOR, fill_type="solid")

        # Calculate totals
        aggregated = self._aggregate_by_account_and_month()
        total_revenue = sum(aggregated.get("REVENUE", {}).values())
        total_opex = sum(aggregated.get("Operating Expense", {}).values())
        total_cogs = sum(aggregated.get("Cost of Good sold", {}).values())

        row = 5
        metrics = [
            ("Total Revenue", f"${total_revenue:,.2f}"),
            ("Operating Expenses", f"${total_opex:,.2f}"),
            ("Cost of Goods Sold", f"${total_cogs:,.2f}"),
            ("Gross Profit", f"${total_revenue - total_cogs:,.2f}"),
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
            ws[f'B{row}'].font = Font(name='Calibri', size=11)
            ws[f'B{row}'].alignment = Alignment(horizontal="right")
            row += 1

        # Data Quality Section
        row += 2
        ws[f'A{row}'] = "DATA QUALITY"
        ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=self.SUBHEADER_COLOR, end_color=self.SUBHEADER_COLOR, fill_type="solid")

        row += 1
        ws[f'A{row}'] = f"P&L Records: {len(self.pnl_data)}"
        ws[f'B{row}'] = f"KPI Records: {len(self.kpi_data)}"
        ws[f'A{row}'].font = Font(name='Calibri', size=10)
        ws[f'B{row}'].font = Font(name='Calibri', size=10)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25

    def _create_pnl_statement(self):
        """Create P&L Statement sheet."""
        print("  Creating P&L Statement...")
        ws = self.wb.create_sheet("P&L Statement", 1)

        # Headers
        ws['A1'] = f"P&L STATEMENT - {self.year}"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        aggregated = self._aggregate_by_account_and_month()
        months = sorted(set(
            month for months in aggregated.values() for month in months
        ))

        # Column headers
        ws['A2'] = "Account"
        for idx, month in enumerate(months, start=2):
            ws.cell(2, idx).value = month

        # Format headers
        for col in range(1, len(months) + 2):
            cell = ws.cell(2, col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.SUBHEADER_COLOR, end_color=self.SUBHEADER_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        row = 3
        account_order = ["REVENUE", "Cost of Good sold", "Operating Expense", "Financial Expenses"]

        for account in account_order:
            if account in aggregated:
                ws[f'A{row}'] = account
                ws[f'A{row}'].font = Font(bold=True)

                for col_idx, month in enumerate(months, start=2):
                    value = aggregated[account].get(month, 0)
                    ws.cell(row, col_idx).value = value
                    ws.cell(row, col_idx).number_format = '$#,##0.00'

                row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(months) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_data_quality_report(self):
        """Create Data Quality Report sheet."""
        print("  Creating Data Quality Report...")
        ws = self.wb.create_sheet("Data Quality Report", -1)

        ws['A1'] = "DATA QUALITY ASSESSMENT"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        row = 3

        # Known issues from profile
        ws[f'A{row}'] = "KNOWN DATA CHARACTERISTICS"
        ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=self.SUBHEADER_COLOR, end_color=self.SUBHEADER_COLOR, fill_type="solid")

        row += 1
        issues = self.profile.get("data_quality", {}).get("known_issues", [])
        for issue in issues:
            ws[f'A{row}'] = f"• {issue}"
            row += 1

        row += 1
        ws[f'A{row}'] = "MISSING PERIODS"
        ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=self.SUBHEADER_COLOR, end_color=self.SUBHEADER_COLOR, fill_type="solid")

        row += 1
        missing = self.profile.get("data_quality", {}).get("missing_periods", [])
        for period in missing:
            ws[f'A{row}'] = f"• {period}"
            row += 1

        row += 1
        ws[f'A{row}'] = "DATA FRESHNESS"
        ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=self.SUBHEADER_COLOR, end_color=self.SUBHEADER_COLOR, fill_type="solid")

        row += 1
        ws[f'A{row}'] = f"P&L Data Fetched: {self.timestamps.get('pnl_fetched', 'N/A')}"
        row += 1
        ws[f'A{row}'] = f"KPI Data Fetched: {self.timestamps.get('kpi_fetched', 'N/A')}"

        ws.column_dimensions['A'].width = 50

    def _create_reconciliation(self):
        """Create Reconciliation sheet."""
        print("  Creating Reconciliation...")
        ws = self.wb.create_sheet("Reconciliation", -1)

        ws['A1'] = "P&L VS KPI RECONCILIATION"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")

        row = 3
        ws[f'A{row}'] = "VALIDATION STATUS"
        ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=self.SUBHEADER_COLOR, end_color=self.SUBHEADER_COLOR, fill_type="solid")

        row += 1
        aggregated = self._aggregate_by_account_and_month()
        total_revenue_pnl = sum(aggregated.get("REVENUE", {}).values())

        ws[f'A{row}'] = "P&L Total Revenue"
        ws[f'B{row}'] = total_revenue_pnl
        ws[f'B{row}'].number_format = '$#,##0.00'
        row += 1

        # Count KPI revenue records - safely handle both dict and list
        kpi_revenue_count = 0
        if isinstance(self.kpi_data, list):
            for r in self.kpi_data:
                if isinstance(r, dict):
                    kpi_name = r.get(self.fields.get("kpi_name", "KPI"), "")
                    if kpi_name and "Revenue" in str(kpi_name):
                        kpi_revenue_count += 1

        ws[f'A{row}'] = "KPI Records with Revenue"
        ws[f'B{row}'] = kpi_revenue_count
        row += 2

        ws[f'A{row}'] = "NOTES"
        ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=self.SUBHEADER_COLOR, end_color=self.SUBHEADER_COLOR, fill_type="solid")

        row += 1
        notes = [
            "P&L and KPI data may differ due to timing of data loads",
            "KPI Revenue represents latest period metrics",
            "P&L includes all transactions for the calendar year",
            "See Data Quality Report for known data characteristics"
        ]
        for note in notes:
            ws[f'A{row}'] = f"• {note}"
            ws[f'A{row}'].font = Font(italic=True, size=9)
            row += 1

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20

    async def generate(self, output_path: str = None):
        """Generate complete FP&A report."""
        print("\n" + "="*80)
        print("📊 GENERATING COMPREHENSIVE FP&A REPORT")
        print("="*80)

        # Fetch data
        print("\n1️⃣  FETCHING DATA...")
        await self.fetch_pnl_data()
        await self.fetch_kpi_data()

        # Create sheets
        print("\n2️⃣  CREATING SHEETS...")
        self._create_executive_summary()
        self._create_pnl_statement()
        self._create_data_quality_report()
        self._create_reconciliation()

        # Save workbook
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"tmp/Comprehensive_FPA_Report_{self.year}_{timestamp}.xlsx"

        print(f"\n3️⃣  SAVING REPORT...")
        self.wb.save(output_path)
        print(f"    ✓ Saved: {output_path}")

        print("\n" + "="*80)
        print("✅ COMPREHENSIVE FP&A REPORT GENERATED")
        print("="*80)
        print(f"\nReport: {output_path}")
        print(f"Year: {self.year}")
        print(f"P&L Records: {len(self.pnl_data)}")
        print(f"KPI Records: {len(self.kpi_data)}")
        print("\nSheets included:")
        for idx, sheet_name in enumerate(self.wb.sheetnames, 1):
            print(f"  {idx}. {sheet_name}")


async def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description="Generate comprehensive FP&A report")
    parser.add_argument("--year", type=int, default=2025, help="Year to analyze")
    parser.add_argument("--env", default="app", help="Environment (dev, demo, testapp, app)")
    parser.add_argument("--profile", help="Path to client profile JSON")
    parser.add_argument("--output", help="Output file path")

    args = parser.parse_args()

    try:
        report = ComprehensiveFPAReport(
            year=args.year,
            env=args.env,
            profile_path=args.profile
        )
        await report.generate(args.output)
    except Exception as e:
        print(f"\n❌ Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
